import openpyxl

# --- 1. Update HTML File ---
html_path = './smoke_test_runner.html'
with open(html_path, 'r', encoding='utf-8') as f:
    html = f.read()

html = html.replace('<title>🏥 UAT Smoke Test Runner & Report Generator</title>', '<title>🏥 Smoke Test Runner & Report Generator</title>')
html = html.replace('UAT Smoke Test Runner & Report Generator', 'Smoke Test Runner & Report Generator')

with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)

print('Updated HTML title successfully.')

# --- 2. Update Excel File ---
excel_path = './E2E flow OPD to IPD Only me.xlsx'
wb = openpyxl.load_workbook(excel_path)

if 'Smoke Test Execution' in wb.sheetnames:
    ws_exec = wb['Smoke Test Execution']
    if 'UAT' in str(ws_exec['A1'].value or ''):
        ws_exec['A1'] = ws_exec['A1'].value.replace('UAT Pre-Verification ', '').replace('UAT ', '')
        print('Updated Excel Smoke Test Execution A1 title.')

if 'Smoke Test Guidelines' in wb.sheetnames:
    ws_guide = wb['Smoke Test Guidelines']
    if 'UAT' in str(ws_guide['A1'].value or ''):
        ws_guide['A1'] = ws_guide['A1'].value.replace('UAT ', '')
        print('Updated Excel Smoke Test Guidelines A1 title.')

wb.save(excel_path)
print('Excel titles updated successfully.')
