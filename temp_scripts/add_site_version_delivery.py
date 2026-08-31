import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- 1. Update Excel File ---
excel_path = './E2E flow OPD to IPD Only me.xlsx'
wb = openpyxl.load_workbook(excel_path)

font_bold = Font(name='Segoe UI', size=10, bold=True, color='000000')
font_normal = Font(name='Segoe UI', size=10, color='000000')
fill_soft_blue = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
thin_side = Side(border_style='thin', color='D9D9D9')
border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

if 'Smoke Test Execution' in wb.sheetnames:
    ws = wb['Smoke Test Execution']
    
    meta_coords = [
        # Row 5
        ("A5", "Hospital Site:", "B5", "รพ.สาธิต / Site Name"),
        ("C5", "System Version:", "D5", "v1.0.0"),
        ("E5", "Delivery Batch (งวดส่งงาน):", "F5", "งวดที่ 1"),
        ("G5", "Test Run ID:", "H5", "SM-RUN-001"),
        
        # Row 6
        ("A6", "Test Cycle:", "B6", "Cycle 1"),
        ("C6", "Date Executed:", "D6", "2026-08-21"),
        ("E6", "Environment:", "F6", "STAGING / UAT"),
        ("G6", "Tester Name:", "H6", "[ชื่อผู้ทดสอบ]"),
        
        # Row 7
        ("A7", "Primary HN:", "B7", "HN......"),
        ("C7", "Primary VN:", "D7", "VN......"),
        ("E7", "Primary AN:", "F7", "AN......"),
        ("G7", "Order IDs:", "H7", "ORD......"),
        
        # Row 8
        ("A8", "Bill No.:", "B8", "INV......"),
        ("C8", "Pass Count:", "D8", "0 / 17"),
        ("E8", "Overall Status:", "F8", "NOT STARTED"),
        ("G8", "", "H8", "")
    ]
    
    for lbl_col, lbl_val, val_col, val_init in meta_coords:
        if lbl_val:
            ws[lbl_col] = lbl_val
            ws[lbl_col].font = font_bold
            ws[lbl_col].alignment = Alignment(horizontal='right', vertical='center')
            ws[lbl_col].fill = fill_soft_blue
            
            ws[val_col] = val_init
            ws[val_col].font = font_normal
            ws[val_col].alignment = Alignment(horizontal='left', vertical='center')
            ws[val_col].border = border_all_thin

    print('Updated Excel Smoke Test Execution header with Site, Version, and Delivery Batch.')

wb.save(excel_path)

# --- 2. Update Web App HTML File ---
html_path = './smoke_test_runner.html'
with open(html_path, 'r', encoding='utf-8') as f:
    html = f.read()

# Add new input fields to Data Chain Form Grid
old_grid_content = """<div class="datachain-grid">
                    <div class="form-group">
                        <label>Run ID</label>
                        <input type="text" id="meta-runid" value="SM-RUN-001" onchange="saveMetadata()">
                    </div>
                    <div class="form-group">
                        <label>Test Cycle</label>
                        <input type="text" id="meta-cycle" value="Cycle 1" placeholder="e.g. Cycle 1, Cycle 2" onchange="saveMetadata()">
                    </div>"""

new_grid_content = """<div class="datachain-grid">
                    <div class="form-group">
                        <label>Hospital Site (ชื่อ รพ.)</label>
                        <input type="text" id="meta-site" value="รพ.สาธิต" placeholder="ชื่อโรงพยาบาล..." onchange="saveMetadata()">
                    </div>
                    <div class="form-group">
                        <label>System Version</label>
                        <input type="text" id="meta-ver" value="v1.0.0" placeholder="v1.0.0" onchange="saveMetadata()">
                    </div>
                    <div class="form-group">
                        <label>Delivery Batch (งวดส่งงาน)</label>
                        <input type="text" id="meta-delivery" value="งวดที่ 1" placeholder="งวดที่ 1..." onchange="saveMetadata()">
                    </div>
                    <div class="form-group">
                        <label>Run ID</label>
                        <input type="text" id="meta-runid" value="SM-RUN-001" onchange="saveMetadata()">
                    </div>
                    <div class="form-group">
                        <label>Test Cycle</label>
                        <input type="text" id="meta-cycle" value="Cycle 1" placeholder="e.g. Cycle 1, Cycle 2" onchange="saveMetadata()">
                    </div>"""

if 'id="meta-site"' not in html:
    html = html.replace(old_grid_content, new_grid_content)

# Update JS save keys list
old_keys = "['runid', 'cycle', 'date', 'tester', 'env', 'hn', 'vn', 'an', 'bill'].forEach"
new_keys = "['site', 'ver', 'delivery', 'runid', 'cycle', 'date', 'tester', 'env', 'hn', 'vn', 'an', 'bill'].forEach"
html = html.replace(old_keys, new_keys)

# Update Linear Report Generator JS text
old_report_js = "let cycle = m.cycle || 'Cycle 1';\n            let text = `📢 **[Smoke Test Summary Report] - Pre-UAT Verification Sign-off**\\\\n`;\n            text += `🗓 Date: ${date} | Test Cycle: ${cycle} | Environment: ${env} | Run ID: ${runid}\\\\n`;"

new_report_js = """let site = m.site || 'รพ.สาธิต';
            let ver = m.ver || 'v1.0.0';
            let delivery = m.delivery || 'งวดที่ 1';
            let cycle = m.cycle || 'Cycle 1';
            let text = `📢 **[Smoke Test Summary Report] - Pre-UAT Verification Sign-off**\\\\n`;
            text += `🏥 Hospital Site: ${site} | Version: ${ver} | Delivery Batch: ${delivery}\\\\n`;
            text += `🗓 Date: ${date} | Test Cycle: ${cycle} | Environment: ${env} | Run ID: ${runid}\\\\n`;"""

html = html.replace(old_report_js, new_report_js)

with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)

print('Updated Web App HTML with Site, Version, and Delivery Batch fields.')
