import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

# --- 1. Update Excel File ---
excel_path = './E2E flow OPD to IPD Only me.xlsx'
wb = openpyxl.load_workbook(excel_path)

if 'Smoke Test Execution' in wb.sheetnames:
    ws = wb['Smoke Test Execution']
    # Set default Hospital Site to NUH
    ws['B5'] = 'NUH (Naresuan University Hospital)'
    
    # Add Data Validation for Site Cell B5
    dv_site = DataValidation(type="list", formula1='"NUH (Naresuan University Hospital), Standard E2E Flow, General Hospital, Custom Site"', allow_blank=True)
    ws.add_data_validation(dv_site)
    dv_site.add("B5")

    print('Updated Excel with NUH as default site and added Site Data Validation dropdown.')

wb.save(excel_path)

# --- 2. Update Web App HTML File ---
html_path = './smoke_test_runner.html'
with open(html_path, 'r', encoding='utf-8') as f:
    html = f.read()

# Replace Hospital Site input with a Select Dropdown + Input combo
old_site_input = """<div class="form-group">
                        <label>Hospital Site (ชื่อ รพ.)</label>
                        <input type="text" id="meta-site" value="รพ.สาธิต" placeholder="ชื่อโรงพยาบาล..." onchange="saveMetadata()">
                    </div>"""

new_site_input = """<div class="form-group">
                        <label>Hospital Site (เลือก Site รพ.)</label>
                        <select id="meta-site-select" onchange="onSiteSelectChange()" style="background: #0F172A; border: 1px solid var(--border-color); color: var(--text-primary); padding: 9px 12px; border-radius: 6px; font-size: 0.9rem; outline: none; cursor: pointer;">
                            <option value="NUH">NUH (Naresuan University Hospital)</option>
                            <option value="Standard">Standard E2E Flow (General Hospital)</option>
                            <option value="Custom">Custom Hospital Site...</option>
                        </select>
                    </div>
                    <div class="form-group" id="custom-site-group" style="display: none;">
                        <label>Custom Site Name</label>
                        <input type="text" id="meta-site" value="NUH (Naresuan University Hospital)" placeholder="ระบุชื่อ รพ..." onchange="saveMetadata()">
                    </div>"""

if 'id="meta-site-select"' not in html:
    html = html.replace(old_site_input, new_site_input)
    html = html.replace('grid-template-columns: repeat(5, 1fr);', 'grid-template-columns: repeat(6, 1fr);')

# Update default values and functions in JS
old_site_init = "let site = m.site || 'รพ.สาธิต';"
new_site_init = "let site = m.site || 'NUH (Naresuan University Hospital)';"

html = html.replace(old_site_init, new_site_init)

# Add onSiteSelectChange JS function
site_js_function = """
        function onSiteSelectChange() {
            let sel = document.getElementById('meta-site-select').value;
            let customGrp = document.getElementById('custom-site-group');
            let inputSite = document.getElementById('meta-site');

            if (sel === 'NUH') {
                customGrp.style.display = 'none';
                inputSite.value = 'NUH (Naresuan University Hospital)';
            } else if (sel === 'Standard') {
                customGrp.style.display = 'none';
                inputSite.value = 'Standard E2E Flow';
            } else {
                customGrp.style.display = 'flex';
                if (inputSite.value.includes('NUH') || inputSite.value.includes('Standard')) {
                    inputSite.value = '';
                }
            }
            saveMetadata();
        }
"""

if 'function onSiteSelectChange()' not in html:
    html = html.replace('function saveMetadata() {', site_js_function + '\n        function saveMetadata() {')

# Adjust JS load state to restore site select
load_state_js = """if (testState.metadata) {
                        for (let k in testState.metadata) {
                            let el = document.getElementById('meta-' + k);
                            if (el && testState.metadata[k]) el.value = testState.metadata[k];
                        }
                    }"""

new_load_state_js = """if (testState.metadata) {
                        for (let k in testState.metadata) {
                            let el = document.getElementById('meta-' + k);
                            if (el && testState.metadata[k]) el.value = testState.metadata[k];
                        }
                        if (testState.metadata.site) {
                            let sel = document.getElementById('meta-site-select');
                            if (testState.metadata.site.includes('NUH')) sel.value = 'NUH';
                            else if (testState.metadata.site.includes('Standard')) sel.value = 'Standard';
                            else {
                                sel.value = 'Custom';
                                document.getElementById('custom-site-group').style.display = 'flex';
                            }
                        }
                    }"""

html = html.replace(load_state_js, new_load_state_js)

with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)

print('Updated Web App HTML with NUH Site Selector.')
