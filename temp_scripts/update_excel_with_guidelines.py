import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('./E2E flow OPD to IPD Only me.xlsx')

# --- 1. Create or Recreate "Smoke Test Guidelines" Sheet ---
guideline_sheet_name = 'Smoke Test Guidelines'
if guideline_sheet_name in wb.sheetnames:
    del wb[guideline_sheet_name]

# Insert as the sheet right before 'Smoke Test Execution' or at end
ws_guide = wb.create_sheet(title=guideline_sheet_name)
ws_guide.views.sheetView[0].showGridLines = True

# --- Styles ---
font_title = Font(name='Segoe UI', size=16, bold=True, color='1F4E79')
font_subtitle = Font(name='Segoe UI', size=11, italic=True, color='595959')
font_sec_header = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
font_card_header = Font(name='Segoe UI', size=11, bold=True, color='1F4E79')
font_bold = Font(name='Segoe UI', size=10, bold=True, color='000000')
font_normal = Font(name='Segoe UI', size=10, color='000000')
font_code = Font(name='Consolas', size=9.5, color='002060')

fill_navy = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
fill_soft_blue = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
fill_light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
fill_yellow_note = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

thin_side = Side(border_style='thin', color='D9D9D9')
thick_side = Side(border_style='medium', color='1F4E79')
border_box = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_side)

align_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
align_left_center = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Title Block
ws_guide['A1'] = "📖 ข้อกำหนดและแนวทางการใช้งาน (Smoke Test Instructions & Guidelines)"
ws_guide['A1'].font = font_title
ws_guide['A2'] = "คำแนะนำสำหรับทีมทดสอบ เพื่อความเข้าใจวัตถุประสงค์ การเก็บหลักฐาน การสรุปผลลง Linear และปัจจัยความสำเร็จ"
ws_guide['A2'].font = font_subtitle

def create_section_header(row, title):
    ws_guide.merge_cells(f'A{row}:G{row}')
    cell = ws_guide[f'A{row}']
    cell.value = title
    cell.font = font_sec_header
    cell.fill = fill_navy
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_guide.row_dimensions[row].height = 28

# SECTION 1: Key Success Factors (วัตถุประสงค์และปัจจัยความสำเร็จ)
create_section_header(4, "🎯 1. วัตถุประสงค์และปัจจัยความสำเร็จ (Key Success Factors)")

ksf_content = (
    "ทำไมต้องทำ Smoke Test ก่อนเปิด UAT?\n"
    "Smoke Test มีเป้าหมายเพื่อ 'ตรวจสอบความพร้อมของระบบในภาพรวม (Quick Verification)' ว่า E2E Flow หลักทำงานได้ ไม่พังทลาย ก่อนที่จะปล่อยให้ผู้ใช้งานจริง (UAT Users) เข้ามาทดสอบ โดยมี 3 ปัจจัยความสำเร็จหลัก ได้แก่:\n\n"
    "1. มี Data Chain Tracker บันทึกรหัสกลางร่วมกัน (Centralized Data Chain):\n"
    "   - ในระบบ HIS ข้อมูลเชื่อมโยงกันข้ามแผนก (HN -> VN -> AN -> Order ID -> Bill No.) ทุกคนในทีมต้องใช้และลงบันทึกในตาราง Header เดียวกัน เพื่อให้เห็นการส่งต่อข้อมูลแบบ End-to-End จริง\n\n"
    "2. โฟกัสการทดสอบแบบ Quick Check / Happy Path:\n"
    "   - เน้นตรวจสอบตามขั้นตอนแบบย่อในตาราง ไม่ต้องจมกับ Edge Cases ในรอบนี้ หากเจอ Bug ติดขัดเกิน 5 นาที ให้ทำเครื่องหมาย FAIL/BLOCKED แล้วข้ามไปข้อถัดไปทันที เพื่อคุมเวลาให้อยู่ใน 45-60 นาที\n\n"
    "3. กำหนด Go / No-Go Decision Gate ชัดเจนก่อนเปิด UAT:\n"
    "   - Scenarios สำคัญระดับ Critical (เช่น SC-01 คัดกรอง, SC-05 จองเตียง, SC-07 IPD Orders, SC-14 Discharge ปิดบิล) ต้อง PASS 100% หากไม่ผ่าน ห้ามเปิดรอบ UAT เด็ดขาด"
)

ws_guide.merge_cells('A5:G5')
cell_ksf = ws_guide['A5']
cell_ksf.value = ksf_content
cell_ksf.font = font_normal
cell_ksf.fill = fill_light_gray
cell_ksf.alignment = align_left_top
cell_ksf.border = border_box
ws_guide.row_dimensions[5].height = 195


# SECTION 2: Evidence & Defect Tracking Guidelines
create_section_header(7, "📸 2. รูปแบบการเก็บผลลัพธ์และหลักฐาน (Evidence & Defect Tracking Guidelines)")

evidence_content = (
    "1. การเก็บหลักฐานภาพถ่ายหน้าจอ (Key Milestone Screenshots):\n"
    "   - ในรอบ Smoke Test ไม่จำเป็นต้องแคปหน้าจอทุกขั้นตอน ให้แคปเฉพาะ '4 จุดสำเร็จสำคัญ' เพื่อยืนยันความสมบูรณ์ของ E2E Flow:\n"
    "     [1] OPD Visit Completed: หน้าจอส่ง Order OPD หรือสร้าง Admit Order สำเร็จ\n"
    "     [2] IPD Admit & Bed Occupied: ผังเตียงที่แสดงผู้ป่วยเข้าพักใน Ward สำเร็จ\n"
    "     [3] MAR / Dispensed: หน้า e-MAR หรือหน้ายืนยันการจ่ายยาสำเร็จ\n"
    "     [4] Discharge & Bill Closed: ใบเสร็จรับเงินยอดสุดท้าย + สถานะจำหน่ายสำเร็จ\n\n"
    "2. มาตรฐานการตั้งชื่อไฟล์ภาพ Bug (Defect Screenshot Naming Convention):\n"
    "   - หากพบข้อผิดพลาด ให้แคปภาพหน้าจอและตั้งชื่อไฟล์ตามรูปแบบมาตรฐาน เพื่อให้ทีม Dev/QA สืบค้นได้ทันที:\n"
    "     รูปแบบ: [ScenarioID]_[Role]_[ShortDescription]_[YYYYMMDD]\n"
    "     ตัวอย่าง: SC05_BedAdmin_CannotGenerateAN_20260821.png\n"
    "             SC09_Pharmacy_StockNotDeducted_20260821.png"
)

ws_guide.merge_cells('A8:G8')
cell_ev = ws_guide['A8']
cell_ev.value = evidence_content
cell_ev.font = font_normal
cell_ev.fill = fill_light_gray
cell_ev.alignment = align_left_top
cell_ev.border = border_box
ws_guide.row_dimensions[8].height = 180


# SECTION 3: Linear Sign-Off Summary Format
create_section_header(10, "📝 3. ตัวอย่างแบบฟอร์ม Dashboard สรุปผล ส่งใน Linear ก่อนเริ่ม UAT")

linear_desc = "เมื่อทดสอบเสร็จเรียบร้อยแล้ว ให้คัดลอก (Copy) ข้อความตามโครงสร้างด้านล่างนี้ ไปโพสต์สรุปผลใน Linear Issue / Requirement การทดสอบ ก่อนแจ้งเปิดรอบ UAT:"
ws_guide.merge_cells('A11:G11')
cell_ld = ws_guide['A11']
cell_ld.value = linear_desc
cell_ld.font = font_bold
ws_guide.row_dimensions[11].height = 22

linear_template_text = (
    "📢 [Smoke Test Summary Report] - Pre-UAT Verification Sign-off\n"
    "🗓 Date: 2026-08-21 | Environment: STAGING / UAT | Run ID: SM-RUN-001\n"
    "👤 Tester: [ระบุชื่อผู้ทดสอบ]\n\n"
    "📊 Overall Status: 🟢 READY FOR UAT (Pass Rate: 94.1%)\n\n"
    "• Total Scenarios: 17 Scenarios\n"
    "• 🟢 Passed: 16 Scenarios\n"
    "• 🔴 Failed: 1 Scenario (SC-15 Report PDF format alignment error - Non-critical)\n"
    "• 🟡 Blocked: 0 Scenarios\n\n"
    "🔗 Primary E2E Data Chain Tracked:\n"
    "• HN: HN69001  |  VN: VN260012  |  AN: AN690088  |  Bill No: INV-551\n\n"
    "⚠️ Open Defect / Issue Links:\n"
    "• [Issue-102] SC-15 Report สรุปยอดตอนเย็น จัดหน้า PDF เหลื่อม (Assignee: DevTeam)\n\n"
    "✅ Recommendation / Sign-off:\n"
    "ระบบผ่านการตรวจความพร้อมหลักครบถ้วน อนุมัติเปิดให้ผู้ใช้งานทดสอบรอบ UAT ได้ตามกำหนดการ"
)

ws_guide.merge_cells('A12:G12')
cell_lt = ws_guide['A12']
cell_lt.value = linear_template_text
cell_lt.font = font_code
cell_lt.fill = fill_yellow_note
cell_lt.alignment = align_left_top
cell_lt.border = border_box
ws_guide.row_dimensions[12].height = 230

# Column widths for Guidelines sheet
guide_col_widths = {'A': 14, 'B': 14, 'C': 16, 'D': 25, 'E': 25, 'F': 25, 'G': 25}
for c_letter, w in guide_col_widths.items():
    ws_guide.column_dimensions[c_letter].width = w

# --- 2. Update Notice Banner in "Smoke Test Execution" Sheet ---
ws_exec = wb['Smoke Test Execution']

# Insert Banner at row 3 (shifting header block clean)
ws_exec.merge_cells('A3:I3')
banner_cell = ws_exec['A3']
banner_cell.value = "⚠️ โปรดอ่านข้อกำหนด รูปแบบการเก็บหลักฐาน ปัจจัยความสำเร็จ และรูปแบบรายงาน Linear ที่ Sheet 'Smoke Test Guidelines' ก่อนเริ่มทำการทดสอบ"
banner_cell.font = Font(name='Segoe UI', size=10, bold=True, color='856404')
banner_cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
banner_cell.alignment = Alignment(horizontal='center', vertical='center')
ws_exec.row_dimensions[3].height = 24

wb.save('./E2E flow OPD to IPD Only me.xlsx')
print('Successfully added "Smoke Test Guidelines" sheet and updated notice banner in Excel!')
