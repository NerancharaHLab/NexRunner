import openpyxl

# --- 1. Update Excel File ---
excel_path = './E2E flow OPD to IPD Only me.xlsx'
wb = openpyxl.load_workbook(excel_path)

if 'Smoke Test Execution' in wb.sheetnames:
    ws = wb['Smoke Test Execution']
    font_common = openpyxl.styles.Font(name='Segoe UI', size=10, bold=True, color='7030A0') # Purple color
    
    for r in range(10, ws.max_row + 1):
        sc_id = str(ws.cell(row=r, column=3).value or '').strip()
        if 'SC-15' in sc_id or 'SC-16' in sc_id:
            cell_flow = ws.cell(row=r, column=2)
            cell_flow.value = 'OPD / IPD'
            cell_flow.font = font_common
            print(f'Updated Row {r} ({sc_id}) flow to "OPD / IPD" in Excel')

    wb.save(excel_path)
    print('Excel updated successfully.')

# --- 2. Update Web App HTML File ---
html_path = './smoke_test_runner.html'
with open(html_path, 'r', encoding='utf-8') as f:
    html = f.read()

# Replace SC-15 flow "IPD" -> "OPD / IPD"
# Replace SC-16 flow "IPD" -> "OPD / IPD"
html = html.replace('id: "SC-15",\n                flow: "IPD",', 'id: "SC-15",\n                flow: "OPD / IPD",')
html = html.replace('id: "SC-16",\n                flow: "IPD",', 'id: "SC-16",\n                flow: "OPD / IPD",')

# Add CSS styling for .flow-badge.opdipd
css_addition = ".flow-badge.opdipd { background: rgba(168, 85, 247, 0.2); color: #C084FC; }"

if '.flow-badge.opdipd' not in html:
    html = html.replace(
        '.flow-badge.ipd { background: rgba(14, 165, 233, 0.2); color: #38BDF8; }',
        '.flow-badge.ipd { background: rgba(14, 165, 233, 0.2); color: #38BDF8; }\n' + css_addition
    )

# Update filter buttons in HTML
old_filter_html = '<div class="filter-buttons">\n                <button class="filter-btn active" onclick="filterCategory(\'all\', this)">ทั้งหมด (17)</button>\n                <button class="filter-btn" onclick="filterCategory(\'OPD\', this)">OPD Flow (4)</button>\n                <button class="filter-btn" onclick="filterCategory(\'IPD\', this)">IPD Flow (13)</button>\n            </div>'

new_filter_html = '<div class="filter-buttons">\n                <button class="filter-btn active" onclick="filterCategory(\'all\', this)">ทั้งหมด (17)</button>\n                <button class="filter-btn" onclick="filterCategory(\'OPD\', this)">OPD Flow (4)</button>\n                <button class="filter-btn" onclick="filterCategory(\'IPD\', this)">IPD Flow (11)</button>\n                <button class="filter-btn" onclick="filterCategory(\'OPD / IPD\', this)">OPD & IPD / General (2)</button>\n            </div>'

html = html.replace(old_filter_html, new_filter_html)

# Update JS badge class name logic
html = html.replace('${sc.flow.toLowerCase()}', '${sc.flow.toLowerCase().replace(/[^a-z0-9]/g, "")}')

# Update JS filter matching logic
old_logic = 'let matchCat = (currentCategory === \'all\') || (flow === currentCategory);'
new_logic = 'let matchCat = (currentCategory === \'all\') || (flow === currentCategory) || (currentCategory !== \'all\' && flow === \'OPD / IPD\');'

html = html.replace(old_logic, new_logic)

with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)

print('Web App HTML updated successfully.')
