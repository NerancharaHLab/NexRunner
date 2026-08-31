import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.load_workbook('./E2E flow OPD to IPD Only me.xlsx')

sheet_name = 'Smoke Test Execution'
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(title=sheet_name)
ws.views.sheetView[0].showGridLines = True

# --- Styles ---
font_title = Font(name='Segoe UI', size=16, bold=True, color='1F4E79')
font_subtitle = Font(name='Segoe UI', size=11, italic=True, color='595959')
font_sec_header = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
font_tbl_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
font_bold = Font(name='Segoe UI', size=10, bold=True, color='000000')
font_normal = Font(name='Segoe UI', size=10, color='000000')
font_opd = Font(name='Segoe UI', size=10, bold=True, color='C00000')
font_ipd = Font(name='Segoe UI', size=10, bold=True, color='1F4E79')

fill_navy = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
fill_soft_blue = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
fill_header_gray = PatternFill(start_color='595959', end_color='595959', fill_type='solid')
fill_zebra = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

thin_side = Side(border_style='thin', color='D9D9D9')
thick_bottom = Side(border_style='medium', color='1F4E79')
double_bottom = Side(border_style='double', color='1F4E79')

border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Title Block
ws['A1'] = "🏥 UAT Pre-Verification Smoke Test Script & Execution Tracker"
ws['A1'].font = font_title
ws['A2'] = "การทดสอบความพร้อมแบบเร่งด่วน (Smoke Test) ครอบคลุม 17 Scenarios ครบถ้วนตาม E2E Flow OPD to IPD"
ws['A2'].font = font_subtitle

# Header Summary Card (Rows 4-7)
ws.merge_cells('A4:I4')
ws['A4'] = "📌 ข้อมูลการทดสอบและสรุปผล (Test Run Information & Data Chain Tracker)"
ws['A4'].font = font_sec_header
ws['A4'].fill = fill_navy
ws['A4'].alignment = Alignment(horizontal='left', vertical='center', indent=1)

info_headers = [
    ("Test Run ID:", "B5"), ("Date Executed:", "D5"), ("Environment:", "F5"), ("Overall Status:", "H5"),
    ("Tester Name:", "B6"), ("Primary HN:", "D6"), ("Primary VN:", "F6"), ("Primary AN:", "H6"),
    ("Order IDs:", "B7"), ("Bill No.:", "D7"), ("Pass Count:", "F7"), ("Fail / Blocked:", "H7")
]

# Set up metadata labels & values placeholders
meta_coords = [
    ("A5", "Test Run ID:", "B5", "SM-RUN-001"),
    ("C5", "Date Executed:", "D5", "2026-08-21"),
    ("E5", "Environment:", "F5", "STAGING / UAT"),
    ("G5", "Overall Status:", "H5", "NOT STARTED"),
    
    ("A6", "Tester Name:", "B6", "[ระบุชื่อผู้ทดสอบ]"),
    ("C6", "Primary HN:", "D6", "HN......"),
    ("E6", "Primary VN:", "F6", "VN......"),
    ("G6", "Primary AN:", "H6", "AN......"),
    
    ("A7", "Order IDs:", "B7", "ORD......"),
    ("C7", "Bill No.:", "D7", "INV......"),
    ("E7", "Pass Count:", "F7", "0 / 17"),
    ("G7", "Fail / Blocked:", "H7", "0")
]

for lbl_col, lbl_val, val_col, val_init in meta_coords:
    ws[lbl_col] = lbl_val
    ws[lbl_col].font = font_bold
    ws[lbl_col].alignment = Alignment(horizontal='right', vertical='center')
    ws[lbl_col].fill = fill_soft_blue
    
    ws[val_col] = val_init
    ws[val_col].font = font_normal
    ws[val_col].alignment = Alignment(horizontal='left', vertical='center')
    ws[val_col].border = border_all_thin

# Table Headers (Row 9)
headers = [
    "ลำดับ\n(No.)", 
    "ระบบ\n(Flow)", 
    "รหัสบททดสอบ\n(Scenario ID)", 
    "ขอบเขตและชื่อบททดสอบ\n(Scenario Scope & Name)", 
    "ผู้ดำเนินการ\n(Role / Actor)", 
    "ขั้นตอนทดสอบแบบย่อ\n(Quick Test Steps)", 
    "เกณฑ์การผ่าน\n(Expected Pass Criteria)", 
    "ผลการทดสอบ\n(Status)", 
    "หมายเหตุ / เลข Bug\n(Notes / Bug ID)"
]

ws.row_dimensions[9].height = 32
for col_num, h_text in enumerate(headers, 1):
    cell = ws.cell(row=9, column=col_num)
    cell.value = h_text
    cell.font = font_tbl_header
    cell.fill = fill_navy
    cell.alignment = align_center
    cell.border = border_header

# Scenario Data
scenarios_data = [
    (1, "OPD", "SC-01", "ลงทะเบียน คัดกรอง ก่อนเข้าตรวจ OPD\n(ผู้ป่วย Walk-in ลงทะเบียนแล้วถูกส่งต่อไปแผนกคัดกรอง และ เข้าคิว OPD)",
     "เวชระเบียน /\nพยาบาลคัดกรอง /\nพยาบาล OPD",
     "1. ค้นหา/สร้าง HN -> ลงทะเบียน Visit -> พิมพ์ใบนำทาง\n2. พยาบาลคัดกรองดู tab 'คำขอ Walk-in' -> บันทึก Vital Sign -> เลือกแพทย์/นัด -> พิมพ์บัตรคิว\n3. คิวขึ้น tab 'รอเช็คอิน' -> กด Check-in -> กดเรียกคิว -> กด 'พร้อมตรวจ'",
     "สถานะผู้ป่วยเปลี่ยนเป็น 'กำลังดำเนินการ' และคิวแสดงหน้าห้องตรวจ"),
    
    (2, "OPD", "SC-02 [A]", "แพทย์ตรวจ + วินิจฉัย + สั่ง Order (Lab/X-ray/ยา)\n(พยาบาลหน้าห้องตรวจกดเรียกคิว ตรวจผู้ป่วย OPD พร้อมกัน เพื่อเข้าห้องตรวจ แล้วสั่ง Order)",
     "แพทย์ OPD /\nพยาบาลหน้าห้องตรวจ",
     "1. พยาบาลกดเรียกคิวเข้าห้องตรวจ\n2. แพทย์เปิด EMR เปลี่ยนสถานะ 'พร้อม' -> บันทึก Diagnosis (ICD-10)\n3. สั่ง Order ครบ 3 ชนิด (Lab + X-ray + ยา) ลงตะกร้า -> กด 'ส่งคำสั่ง' (Submit Order)",
     "ส่ง Order ทั้งหมดสำเร็จในครั้งเดียว และสถานะผู้ป่วยเปลี่ยนเป็น 'พักการตรวจ'"),
    
    (3, "OPD", "SC-03", "ส่ง Lab/X-ray -> รับผล -> แพทย์อ่านผลต่อ\n(ระบบส่ง Order ไปห้อง Lab/X-ray และประมวลผล Lab/X-ray พร้อมแพทย์อ่านผลกลับเข้าตรวจต่อ)",
     "จนท. Lab /\nแพทย์ OPD",
     "1. จนท. Lab เช็คอินรายการ -> เก็บตัวอย่าง (Container) -> ยืนยัน -> ลงผล Lab -> กด Final Report\n2. แพทย์เปิด EMR ดูผล -> เปลี่ยนสถานะ 'พร้อม' -> สั่งยาเพิ่ม/นัดหมาย/ออกใบรับรองแพทย์ -> กด 'ตรวจเสร็จ'",
     "ผล Lab ออกสมบูรณ์ และแพทย์เปลี่ยนสถานะผู้ป่วยเป็น 'ตรวจเสร็จ'"),
    
    (4, "OPD", "SC-04", "เภสัชกรรม (สร้างใบยา) -> ชำระเงิน -> เภสัชกรรม จ่ายยา -> การเงินปิด Visit\n(เภสัชกรสร้างใบยาส่งต่อการเงินชำระเงินแล้วเภสัชกรจ่ายยาปิด Visit ให้ผู้ป่วย OPD)",
     "เภสัชกร OPD /\nการเงิน",
     "1. เภสัชกรตรวจใบสั่งยา -> กด 'สร้างใบยา' -> กด 'ยืนยันใบยา' (ตัด Stock)\n2. การเงินออกใบแจ้งหนี้ -> รับชำระเงิน -> พิมพ์ใบเสร็จ -> กด 'เสร็จสิ้น'\n3. เภสัชกรเห็นสถานะ 'ชำระเงินแล้ว' -> พิมพ์ฉลากยา/จัดยา -> กด 'จ่ายยา' + เรียกคิว",
     "ตัด Stock ยาสำเร็จ ออกใบเสร็จเรียบร้อย และปิด Visit OPD สมบูรณ์"),
    
    (5, "IPD", "SC-02 [B]", "แพทย์วินิจฉัยแล้วสั่ง Admit โดยไม่มี Order\n(แพทย์วินิจฉัยแล้วสั่ง Admit ผู้ป่วย OPD เข้า IPD ทันทีโดยไม่สั่ง Order)",
     "แพทย์ OPD /\nพยาบาล",
     "1. พยาบาลเรียกคิว -> แพทย์เปิด EMR เปลี่ยนสถานะ 'พร้อม' -> ลงวินิจฉัย OPD Note\n2. แพทย์ประเมิน 'ต้อง Admit' -> กดปุ่ม 'สร้าง Admit Order' จาก EMR -> กรอกฟอร์ม -> Save",
     "สถานะผู้ป่วยเปลี่ยนเป็น 'รอ Admit' และคำขอส่งไปศูนย์จัดการเตียง"),
    
    (6, "IPD", "SC-05", "จองเตียง -> สร้าง AN (ต่อเนื่องจาก Scenario 2 [B])\n(จนท.เตียงจองเตียงและสร้าง AN ให้ผู้ป่วย Admit)",
     "จนท. ศูนย์จัดการเตียง",
     "1. เข้าเมนู ศูนย์จัดการเตียง -> tab 'คำขอ Admit' -> ค้นหาผู้ป่วย -> กด 'จองเตียง'\n2. เลือกวอร์ด/ประเภทเตียง/เตียง จากผังเตียง -> กดบันทึก\n3. กลับ tab 'คำขอ Admit' -> กดปุ่ม 'สร้าง AN'",
     "ระบบ Generate AN สำเร็จ ข้อมูลย้ายไปแสดงใน tab 'AN'"),
    
    (7, "IPD", "SC-06", "ผู้ป่วยใน -> รายการเตียง หอผู้ป่วยใน\n(พยาบาล Ward เปิดรายการเตียงหอผู้ป่วยในและบันทึกเวลาผู้ป่วยมาถึงวอร์ด)",
     "พยาบาล Ward",
     "1. เข้าเมนู ผู้ป่วยใน -> รายการเตียง -> เลือก Ward -> กด ส่ง\n2. ไปที่ tab 'เตียงทั้งหมด' -> เลือกผู้ป่วย -> กด 3 จุด -> เลือก 'บันทึกมาถึงวอร์ด' -> บันทึกเวลา",
     "บันทึกเวลาผู้ป่วยมาถึงวอร์ดสำเร็จ ผู้ป่วยแสดงสถานะพร้อมรับการดูแลบน Ward"),
    
    (8, "IPD", "SC-07", "แพทย์ตรวจ + วินิจฉัย + สั่ง Order (Lab/X-ray/ยา)\n(แพทย์ Round ตรวจและสั่ง Order ยา/Lab/X-ray แบบ ONE DAY และ CONTINUE ให้ผู้ป่วย IPD)",
     "แพทย์ IPD",
     "1. เปิด EMR IPD (ค้นด้วย AN) -> เข้าเมนู 'IPD Summary' -> บันทึก Progress Note\n2. เข้า tab 'Physician Order' -> สั่งออเดอร์ทั้ง 'One Day Order' และ 'Continue Order'\n3. กด บันทึก -> กดปุ่ม 'Sign All'",
     "Order แสดงใน tab IPD Summary แยกกลุ่ม One Day / Continue ชัดเจน Status = 'รอรับ'"),
    
    (9, "IPD", "SC-08", "พยาบาลตรวจสอบ Order ยา และ Confirm รายการยา\n(พยาบาลตรวจสอบและ Confirm รายการยาผู้ป่วย IPD แพทย์สั่ง Order ทั้งแบบ ONE DAY และ CONTINUE)",
     "พยาบาล Ward",
     "1. เข้าเมนู ผู้ป่วยใน -> รายการเตียง -> เลือก Ward -> เปิดหน้า 'Patient IPD Summary'\n2. ตรวจสอบคอลัมน์ 'ORDER FOR ONE DAY' และ 'NEW ORDER FOR CONTINUE'\n3. ตรวจสอบชื่อยา/ขนาด/วิธีใช้ -> กดปุ่ม 'รับออเดอร์' ทุกรายการ",
     "สถานะ Order เปลี่ยนเป็นรับออเดอร์แล้ว และรายการถูกส่งไปห้องยา IPD"),
    
    (10, "IPD", "SC-09", "เภสัชกรรมจ่ายยา IPD [ONE DAY] (สร้างใบยา) -> ชำระเงิน -> เภสัชกรรม จ่ายยา -> พยาบาล Ward\n(เภสัชกรจ่ายยาผู้ป่วย IPD พร้อมกัน X ราย แยกตามประเภท Order ONE DAY และต้องชำระเงิน)",
     "เภสัชกร IPD /\nการเงิน /\nพยาบาล Ward",
     "1. เภสัชกรเข้า tab 'Order รอยืนยัน' -> ตรวจสอบยา One Day -> กด 'สร้างใบยา' (พิมพ์ฉลาก auto) -> ยืนยันใบยา\n2. การเงินออกใบแจ้งหนี้ -> รับชำระเงิน\n3. เภสัชกรเห็นสถานะ 'ชำระเงินแล้ว' -> จัดยา -> กด 'จ่ายยา'",
     "รายการยา One Day วิ่งไปแสดงบน e-MAR ของพยาบาล Ward พร้อมจ่ายยา"),
    
    (11, "IPD", "SC-10", "เภสัชกรรมจ่ายยา IPD [CONTINUE] (สร้างใบยา) -> เภสัชกรรม จ่ายยา -> พยาบาล Ward\n(เภสัชกรจ่ายยาผู้ป่วย IPD พร้อมกัน X ราย แยกตามประเภท CONTINUE โดยไม่รอการชำระเงิน)",
     "เภสัชกร IPD /\nพยาบาล Ward",
     "1. เภสัชกรเข้าเมนู 'รับ Continue Order ประจำวัน' -> เลือก Ward -> กด ค้นหา\n2. กดปุ่ม 'Generate ที่เลือก' -> ระบบสร้างใบยา Continue\n3. พิมพ์ฉลากยา -> กด 'จ่ายยา' (ไม่ต้องรอชำระเงิน)",
     "รายการยา Continue ประจำวันถูก Generate และแสดงบน e-MAR ของ Ward ทันที"),
    
    (12, "IPD", "SC-11", "จนท. Lab check-in + ลงผล + แพทย์อ่านผล\n(จนท. Lab ดำเนินการตรวจและบันทึกผลผู้ป่วย IPD ตามประเภท Order ONE DAY)",
     "จนท. Lab /\nแพทย์ IPD",
     "1. จนท. Lab เข้า รายการแล็บ -> tab 'รอเช็คอิน' -> กด เช็คอิน -> เลือกประเภทตัวอย่าง -> กด เก็บแล้ว (Container) -> กด รับแล้ว\n2. ลงผล Lab -> บันทึกร่าง -> กด Final Report\n3. แพทย์เปิด Patient Profile IPD -> เมนู Lab result เพื่อดูผล",
     "สถานะ Lab Task = 'เสร็จสิ้น' (Result completed) และแพทย์เห็นผลในระบบ"),
    
    (13, "IPD", "SC-12", "จนท. Xray check-in + ลงผล + แพทย์อ่านผล\n(จนท. Xray ดำเนินการตรวจและบันทึกผลผู้ป่วย IPD พร้อมกัน X ราย แยกตามประเภท Order ONE DAY)",
     "จนท. X-ray /\nพยาบาล /\nแพทย์อ่านผล",
     "1. จนท. X-ray ไป tab 'รอ Check-in' -> เช็คอิน -> กด 'เริ่มถ่ายภาพ' -> กด 'ถ่ายภาพแล้ว' -> เลือกแพทย์อ่านผล\n2. พยาบาลลงบันทึก Nurse Note / Imaging Nurse Note\n3. แพทย์ลงรายงานผลตรวจรังสี -> กด บันทึก",
     "สถานะ Log = อ่านผล และรายงานผลรังสีแสดงใน EMR สมบูรณ์"),
    
    (14, "IPD", "SC-13", "การเงินเรียกเก็บระหว่างพักรักษา (Interim Billing)\n(ญาติผู้ป่วยชำระค่ารักษาสะสมของผู้ป่วย IPD ทุก 3 วัน ตลอดช่วงพักรักษาในโรงพยาบาล)",
     "การเงิน (Cashier IPD)",
     "1. เข้าเมนู Cashier ค้นหาผู้ป่วย IPD -> ดูรายการค่าใช้จ่ายสะสม (ยา/Lab/X-ray/ค่าห้อง)\n2. ตรวจสอบสิทธิ -> เลือกรายการค่าใช้จ่าย -> กด 'สร้างใบแจ้งหนี้' (Interim Bill)\n3. เลือกสิทธิ/วิธีชำระ -> กด จ่าย + ยืนยัน",
     "ระบบหักยอดชำระแล้วออกจากยอดค้างสะสม และตั้งรอบบันทึกใหม่สำหรับรอบถัดไป"),
    
    (15, "IPD", "SC-14", "Discharge Process (ปิดยอดครั้งสุดท้าย)\n(แพทย์อนุมัติจำหน่ายผู้ป่วย IPD แล้วจนท.การเงินปิด Bill ยอดคงเหลือทั้งหมดและปลดเตียงคืนระบบ)",
     "แพทย์ IPD /\nการเงิน IPD",
     "1. แพทย์บันทึก Discharge Summary + ลงวินิจฉัยสุดท้าย + สั่งยากลับบ้าน -> กด ยืนยัน Discharge\n2. การเงินเปิดรายการ IPD รอปิด Bill -> ออกใบแจ้งหนี้ยอดสุดท้าย -> กด จ่าย + ยืนยัน -> ออกใบเสร็จ\n3. การเงินกดปุ่ม 'ปิด Admission'",
     "รายชื่อผู้ป่วยย้ายไป tab 'ผู้ป่วยจำหน่าย' (Status = เสร็จสิ้น) และปลดเตียงคืนระบบสำเร็จ"),
    
    (16, "IPD", "SC-15", "จนท. จากทุกแผนกเข้าสั่งพิมพ์ report จากเมนู รายงาน\n(เจ้าหน้าที่แต่ละแผนกละคน เข้ามาออกรายงานที่เมนูออกรายงาน)",
     "จนท. ทุกแผนก",
     "1. เข้าเมนู รายงาน (Report) -> เลือกประเภทรายงานตามแผนก\n2. ทดสอบกด สั่งพิมพ์ (Print) สรุปยอด\n3. ทดสอบกด ดาวน์โหลด (Download) สรุปยอด",
     "สั่งออกรายงานได้สำเร็จ ข้อมูลในไฟล์ PDF/Excel ถูกต้องครบถ้วน"),
    
    (17, "IPD", "SC-16", "จนท. Coder เข้าเมนู เบิกจ่าย เพื่อลงรหัสโรค\n(เจ้าหน้าที่แผนก Coder เข้ามาลงรหัสโรคของผู้ป่วย)",
     "จนท. Coder",
     "1. เข้าเมนู เบิกจ่าย -> OPD/IPD -> รายการรอรหัสโรค -> เลือกผู้ป่วยสถานะ 'รอลงรหัส'\n2. ไปที่ Patient Profile -> เลือก Diagnosis + Procedure -> กด บันทึก\n3. กดปุ่ม '>' เพื่อลงรหัสคนถัดไป",
     "สถานะผู้ป่วยเปลี่ยนเป็น 'ลงรหัสแล้ว' และย้ายไปแสดงใน tab ลงรหัสแล้ว")
]

start_row = 10
for i, sc in enumerate(scenarios_data):
    r = start_row + i
    ws.row_dimensions[r].height = 55
    
    no_val, flow_val, sc_id, sc_name, role_val, steps_val, pass_val = sc
    
    ws.cell(row=r, column=1, value=no_val).alignment = align_center
    
    cell_flow = ws.cell(row=r, column=2, value=flow_val)
    cell_flow.alignment = align_center
    cell_flow.font = font_opd if flow_val == "OPD" else font_ipd
    
    cell_sc_id = ws.cell(row=r, column=3, value=sc_id)
    cell_sc_id.alignment = align_center
    cell_sc_id.font = font_bold
    
    ws.cell(row=r, column=4, value=sc_name).alignment = align_top_left
    ws.cell(row=r, column=5, value=role_val).alignment = align_top_left
    ws.cell(row=r, column=6, value=steps_val).alignment = align_top_left
    ws.cell(row=r, column=7, value=pass_val).alignment = align_top_left
    
    # Status column (Col H = 8)
    cell_status = ws.cell(row=r, column=8, value="Not Run")
    cell_status.alignment = align_center
    cell_status.font = font_bold
    
    # Notes column (Col I = 9)
    ws.cell(row=r, column=9, value="").alignment = align_top_left
    
    # Apply fonts & borders
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c)
        cell.border = border_all_thin
        if c not in [2, 3, 8]:
            cell.font = font_normal
        if i % 2 == 1:
            cell.fill = fill_zebra

# Data Validation for Status Column (Column H)
dv = DataValidation(type="list", formula1='"Not Run, Passed, Failed, Blocked"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"H10:H{start_row + len(scenarios_data) - 1}")

# Set Column Widths
col_widths = {
    'A': 8,   # No
    'B': 10,  # Flow
    'C': 16,  # Scenario ID
    'D': 35,  # Scope & Name
    'E': 20,  # Role
    'F': 45,  # Steps
    'G': 35,  # Pass Criteria
    'H': 15,  # Status
    'I': 25   # Notes / Bug ID
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

wb.save('./E2E flow OPD to IPD Only me.xlsx')
print('Successfully created and styled "Smoke Test Execution" sheet in Excel!')
