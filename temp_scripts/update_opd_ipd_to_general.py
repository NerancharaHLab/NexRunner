# -*- coding: utf-8 -*-
import glob
import json
import openpyxl

# --- 1. Update all JSON scenario files in ./scenarios/*.json ---
for fpath in glob.glob('./scenarios/*.json'):
    with open(fpath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    for sc in data.get('scenarios', []):
        if sc.get('flow') == 'OPD / IPD':
            sc['flow'] = 'General'
            
    with open(fpath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

print('Updated JSON scenario files: flow changed to General for SC-15 and SC-16.')

# --- 2. Update HTML Files ---
for filename in ['index.html', 'smoke_test_runner.html']:
    with open(f'./{filename}', 'r', encoding='utf-8') as f:
        html = f.read()
    
    # Replace Filter Button HTML
    html = html.replace(
        '<button class="filter-btn" onclick="filterCategory(\'OPD / IPD\', this)">OPD / IPD (2)</button>',
        '<button class="filter-btn" onclick="filterCategory(\'General\', this)">General (2)</button>'
    )
    
    with open(f'./{filename}', 'w', encoding='utf-8') as f:
        f.write(html)

print('Updated HTML filter buttons.')

# --- 3. Update CSS Stylesheet ---
with open('./css/styles.css', 'r', encoding='utf-8') as f:
    css = f.read()

css = css.replace(
    '.flow-badge.opdipd { background: rgba(168, 85, 247, 0.2); color: #C084FC; }',
    '.flow-badge.general { background: rgba(168, 85, 247, 0.2); color: #C084FC; }'
)

with open('./css/styles.css', 'w', encoding='utf-8') as f:
    f.write(css)

print('Updated CSS flow badge class.')

# --- 4. Update JS logic in app.js ---
with open('./js/app.js', 'r', encoding='utf-8') as f:
    js_code = f.read()

js_code = js_code.replace(
    'let matchCat = (currentCategory === \'all\') || (flow === currentCategory) || (currentCategory !== \'all\' && flow === \'OPD / IPD\');',
    'let matchCat = (currentCategory === \'all\') || (flow === currentCategory);'
)

# Update fallback data flows as well
js_code = js_code.replace('flow: "OPD / IPD"', 'flow: "General"')

with open('./js/app.js', 'w', encoding='utf-8') as f:
    f.write(js_code)

print('Updated JS filter logic in app.js.')

# --- 5. Update Excel File ---
excel_path = './E2E flow OPD to IPD Only me.xlsx'
wb = openpyxl.load_workbook(excel_path)

if 'Smoke Test Execution' in wb.sheetnames:
    ws = wb['Smoke Test Execution']
    for row in range(12, ws.max_row + 1):
        sc_id = str(ws.cell(row=row, column=1).value or '')
        if 'SC-15' in sc_id or 'SC-16' in sc_id:
            ws.cell(row=row, column=2).value = 'General'

    print('Updated Excel Sheet Smoke Test Execution (SC-15 and SC-16 flow = General).')

wb.save(excel_path)
