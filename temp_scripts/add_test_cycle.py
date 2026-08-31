import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- 1. Update Excel File ---
excel_path = './E2E flow OPD to IPD Only me.xlsx'
wb = openpyxl.load_workbook(excel_path)

font_bold = Font(name='Segoe UI', size=10, bold=True, color='000000')
font_normal = Font(name='Segoe UI', size=10, color='000000')
fill_soft_blue = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
thin_side = Side(border_style='thin', color='D9D9D9')
border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

if 'Smoke Test Execution' in wb.sheetnames:
    ws = wb['Smoke Test Execution']
    
    # Adjust Row 5-7 grid to include Test Cycle
    # A5: Test Run ID | C5: Test Cycle | E5: Date Executed | G5: Environment / Status
    meta_coords = [
        ("A5", "Test Run ID:", "B5", "SM-RUN-001"),
        ("C5", "Test Cycle:", "D5", "Cycle 1"),
        ("E5", "Date Executed:", "F5", "2026-08-21"),
        ("G5", "Environment:", "H5", "STAGING / UAT"),
        
        ("A6", "Tester Name:", "B6", "[ชื่อผู้ทดสอบ]"),
        ("C6", "Primary HN:", "D6", "HN......"),
        ("E6", "Primary VN:", "F6", "VN......"),
        ("G6", "Primary AN:", "H6", "AN......"),
        
        ("A7", "Order IDs:", "B7", "ORD......"),
        ("C7", "Bill No.:", "D7", "INV......"),
        ("E7", "Pass Count:", "F7", "0 / 17"),
        ("G7", "Overall Status:", "H7", "NOT STARTED")
    ]
    
    for lbl_col, lbl_val, val_col, val_init in meta_coords:
        ws[lbl_col] = lbl_val
        ws[lbl_col].font = font_bold
        ws[lbl_col].alignment = Alignment(horizontal='right', vertical='center')
        ws[lbl_col].fill = fill_soft_blue
        
        ws[val_col] = val_init
        ws[val_col].font = font_normal
        ws[val_col].alignment = Alignment(horizontal='left', vertical='center')
        ws[val_col].border = border_all_thin

    print('Updated Smoke Test Execution sheet with Test Cycle field.')

wb.save(excel_path)

# --- 2. Update Web App HTML ---
html_path = './smoke_test_runner.html'
with open(html_path, 'r', encoding='utf-8') as f:
    html = f.read()

# Add Test Cycle field to form grid
old_form_grid = """<div class="datachain-grid">
                    <div class="form-group">
                        <label>Run ID</label>
                        <input type="text" id="meta-runid" value="SM-RUN-001" onchange="saveMetadata()">
                    </div>"""

new_form_grid = """<div class="datachain-grid">
                    <div class="form-group">
                        <label>Run ID</label>
                        <input type="text" id="meta-runid" value="SM-RUN-001" onchange="saveMetadata()">
                    </div>
                    <div class="form-group">
                        <label>Test Cycle</label>
                        <input type="text" id="meta-cycle" value="Cycle 1" placeholder="e.g. Cycle 1, Cycle 2" onchange="saveMetadata()">
                    </div>"""

if 'id="meta-cycle"' not in html:
    html = html.replace(old_form_grid, new_form_grid)
    # Adjust grid columns from 4 to 5 if needed, or flex
    html = html.replace('grid-template-columns: repeat(4, 1fr);', 'grid-template-columns: repeat(5, 1fr);')

# Save metadata update in JS
old_js_save = "['runid', 'date', 'tester', 'env', 'hn', 'vn', 'an', 'bill'].forEach"
new_js_save = "['runid', 'cycle', 'date', 'tester', 'env', 'hn', 'vn', 'an', 'bill'].forEach"
html = html.replace(old_js_save, new_js_save)

# Linear Report string update in JS
old_report_line = "let text = `📢 **[Smoke Test Summary Report] - Pre-UAT Verification Sign-off**\\\\n`;\n            text += `🗓 Date: ${date} | Environment: ${env} | Run ID: ${runid}\\\\n`;"

new_report_line = "let cycle = m.cycle || 'Cycle 1';\n            let text = `📢 **[Smoke Test Summary Report] - Pre-UAT Verification Sign-off**\\\\n`;\n            text += `🗓 Date: ${date} | Test Cycle: ${cycle} | Environment: ${env} | Run ID: ${runid}\\\\n`;"

html = html.replace(old_report_line, new_report_line)

with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)

print('Updated Web App HTML with Test Cycle field.')
